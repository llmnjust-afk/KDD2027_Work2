"""DSBench real-task adapter for external validity (v2: xlsx-native).

FIX v2: The original v1 converted xlsx -> csv, which destroyed multi-sheet
structure and caused 0% success rate. This version keeps the ORIGINAL .xlsx
file and tells the agent to use pd.read_excel(), matching how DSBench's own
agent works. This is critical -- the 0% success rate in v1 was a data-pipeline
bug, not a real finding.

Also pre-inspects each xlsx to extract sheet names and a column preview, which
is injected into the task description so the agent knows the data layout.
"""

from __future__ import annotations

import ast
import io
import os
import zipfile
from typing import List, Optional, Tuple
from .tasks import Task, TaskSet, Trap


def _load_dsbench_index(data_json_path: str) -> list:
    """Load DSBench data.json (Python-dict-per-line format)."""
    samples = []
    with open(data_json_path) as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    samples.append(ast.literal_eval(line))
                except Exception:
                    continue
    return samples


def _inspect_xlsx(xlsx_bytes: bytes) -> str:
    """Extract sheet names and a column preview to help the agent.

    Returns a short text description like:
      'Sheets: [Inputs, Calculations, Outputs]. First sheet columns: [A, B, C]'
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
        sheet_names = wb.sheetnames[:5]  # top 5 sheets
        # preview first sheet's first few non-empty rows
        ws = wb[wb.sheetnames[0]]
        preview_rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 8:
                break
            vals = [str(c)[:15] if c is not None else "" for c in row[:8]]
            if any(vals):
                preview_rows.append(vals)
        wb.close()
        desc = f"Sheets: {sheet_names}. "
        if preview_rows:
            desc += f"First sheet preview (first {len(preview_rows)} rows): "
            desc += " | ".join(",".join(r) for r in preview_rows[:3])
        return desc[:500]
    except Exception:
        return "Multi-sheet Excel workbook (use pd.read_excel to inspect)"


def _extract_task_files(zip_path: str, task_id: str) -> Tuple[Optional[bytes], List[str], Optional[str], str]:
    """Extract xlsx bytes, question texts, intro, and xlsx description.

    Returns (xlsx_bytes, question_texts, intro_text, xlsx_description).
    """
    with zipfile.ZipFile(zip_path) as z:
        prefix = f"data/{task_id}/"
        names = [n for n in z.namelist() if n.startswith(prefix) and not n.startswith("__MACOSX")]

        # find xlsx
        xlsx_name = next((n for n in names if n.endswith(".xlsx")), None)
        xlsx_bytes = None
        xlsx_desc = ""
        if xlsx_name:
            xlsx_bytes = z.read(xlsx_name)
            xlsx_desc = _inspect_xlsx(xlsx_bytes)

        # find questions
        q_names = sorted([n for n in names if "question" in n.lower() and n.endswith(".txt")])
        question_texts = []
        for qn in q_names:
            txt = z.read(qn).decode("utf-8", errors="replace").strip()
            question_texts.append(txt)

        # find intro
        intro_name = next((n for n in names if "intro" in n.lower() and n.endswith(".txt")), None)
        intro = z.read(intro_name).decode("utf-8", errors="replace").strip() if intro_name else ""

        return xlsx_bytes, question_texts, intro, xlsx_desc


def _make_checker(answer: str):
    """Build an answer checker for a DSBench answer."""
    ans_str = str(answer).strip()
    ans_upper = ans_str.upper()

    # MC: single letter
    if ans_upper in "ABCDEFGHIJ":
        def _mc(pred, _gold, _a=ans_upper):
            if pred is None:
                return False
            return _a in str(pred).strip().upper()[:5]
        return _mc

    # numeric
    try:
        ans_float = float(ans_str)
        def _num(pred, _gold, _a=ans_float):
            if pred is None:
                return False
            try:
                return abs(float(pred) - _a) <= abs(_a) * 0.02 + 1  # 2% tolerance
            except (ValueError, TypeError):
                return str(_a) in str(pred).replace(",", "")
        return _num
    except ValueError:
        pass

    # text: substring match
    def _text(pred, _gold, _a=ans_str.lower()):
        if pred is None:
            return False
        return _a in str(pred).lower()
    return _text


def _make_data_gen(xlsx_bytes: bytes, xlsx_filename: str = "data.xlsx"):
    """Create a data_generator that writes the ORIGINAL .xlsx file."""
    def _gen(workdir: str, _xlsx=xlsx_bytes, _fname=xlsx_filename):
        os.makedirs(workdir, exist_ok=True)
        with open(os.path.join(workdir, _fname), "wb") as f:
            f.write(_xlsx)
    return _gen


def build_dsbench_real_subset(
    dsbench_repo_path: str = "/data/lab/DSBench",
    n_questions: int = 200,
    preferred_task_ids: Optional[list] = None,
) -> TaskSet:
    """Build a TaskSet from REAL DSBench tasks (xlsx-native, v2).

    Each task ships the original .xlsx file and instructs the agent to use
    pd.read_excel(), fixing the 0%-success-rate bug from v1's csv conversion.

    Parameters
    ----------
    preferred_task_ids : list, optional
        If given, these DSBench task IDs (e.g. ['00000043','00000033']) are
        processed first, so simpler tasks populate the subset before harder
        ones. This avoids spending the n_questions budget on the first
        (hardest) multi-question challenge.
    """
    data_json = os.path.join(dsbench_repo_path, "data_analysis", "data.json")
    data_zip = os.path.join(dsbench_repo_path, "data_analysis", "data_old.zip")

    if not os.path.exists(data_json):
        raise FileNotFoundError(f"DSBench data.json not found at {data_json}")
    if not os.path.exists(data_zip):
        raise FileNotFoundError(f"DSBench data_old.zip not found at {data_zip}")

    samples = _load_dsbench_index(data_json)

    # reorder: preferred (simple) tasks first
    if preferred_task_ids:
        pref = [s for s in samples if s.get("id") in preferred_task_ids]
        rest = [s for s in samples if s.get("id") not in preferred_task_ids]
        samples = pref + rest

    tasks: List[Task] = []
    count = 0

    for sample in samples:
        if count >= n_questions:
            break
        task_id_raw = sample.get("id", "")
        questions = sample.get("questions", [])
        answers = sample.get("answers", [])
        task_name = sample.get("name", "unknown")

        if not questions:
            continue

        # extract real data files for this task
        try:
            xlsx_bytes, q_texts, intro, xlsx_desc = _extract_task_files(data_zip, task_id_raw)
        except Exception:
            continue

        if xlsx_bytes is None:
            continue  # skip tasks without xlsx

        for qi, (qname, ans) in enumerate(zip(questions, answers)):
            if count >= n_questions:
                break

            # get question text
            q_text = q_texts[qi] if qi < len(q_texts) else f"Question: {qname}"

            # build full question with context -- tell agent to use read_excel
            intro_snippet = intro[:1200] if intro else ""
            full_question = (
                f"(DSBench Real Task) {task_name}\n\n"
                f"Context: {intro_snippet}\n\n"
                f"Data file info: {xlsx_desc}\n\n"
                f"Question: {q_text}\n\n"
                f"The data is in 'data.xlsx' (an Excel file with multiple sheets). "
                f"Load it with: df = pd.read_excel('data.xlsx', sheet_name=None) "
                f"to get all sheets as a dict. Then analyze and answer. "
                f"Print 'ANSWER: <your answer>'"
            )

            ans_str = str(ans).strip()
            task_id = f"dsbench_real_{task_id_raw}_{qname}"

            tasks.append(Task(
                task_id=task_id,
                domain="dsbench_real",
                question=full_question,
                answer=ans_str,
                gt_path=["read_excel data.xlsx", "identify correct sheet", "analyze financial data", "compute answer"],
                traps=[
                    Trap("misread_financial_data", "interpretation",
                         "incorrect financial formula or data interpretation", is_silent=True),
                    Trap("wrong_sheet", "execution",
                         "reads wrong sheet of the multi-sheet workbook", is_silent=False),
                ],
                data_generator=_make_data_gen(xlsx_bytes),
                answer_checker=_make_checker(ans_str),
                metadata={
                    "source": "dsbench_real",
                    "original_id": task_id_raw,
                    "original_name": task_name,
                    "year": sample.get("year", ""),
                    "answer_type": "MC" if ans_str.upper() in "ABCDEFGHIJ"
                                  else "numeric" if ans_str.replace(".", "").replace("-", "").isdigit()
                                  else "text",
                },
            ))
            count += 1

    print(f"Built {len(tasks)} real DSBench tasks (xlsx-native v2)")
    return TaskSet(tasks=tasks)
